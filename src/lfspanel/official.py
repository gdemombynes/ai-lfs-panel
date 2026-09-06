"""Fetch published headline rates used to validate the harmonization.

Each fetcher returns rows ``period, indicator, value, tolerance, population,
source_url`` for the indicators produced by ``lfspanel.validate.headline_rates``.
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Dict, Iterable, List, Optional

import openpyxl
import pandas as pd
import requests

from lfspanel.config import RAW
from lfspanel.fetch.base import download, make_session, url_exists
from lfspanel.periods import Period


# ---------------------------------------------------------------- shared
def _rows(
    period: Period,
    url: str,
    total: float,
    lf: float,
    emp: float,
    unemp: float,
    tol: float,
    pop: str,
) -> List[dict]:
    base = {
        "period": str(period),
        "population": pop,
        "source_url": url,
        "tolerance": tol,
    }
    return [
        dict(base, indicator="participation_rate", value=round(100 * lf / total, 3)),
        dict(base, indicator="unemployment_rate", value=round(100 * unemp / lf, 3)),
        dict(base, indicator="employment_rate", value=round(100 * emp / total, 3)),
        dict(base, indicator="population", value=total, tolerance=0.005 * total),
        dict(base, indicator="employed", value=emp, tolerance=0.005 * emp),
    ]


# ---------------------------------------------------------------- Brazil
SIDRA = (
    "https://apisidra.ibge.gov.br/values/t/4092/n1/all/v/all/p/{period}/c629/all"
    "?formato=json"
)
SIDRA_TOL = 0.06  # published rates are rounded to one decimal


def sidra_bra(
    periods: Iterable[Period], session: Optional[requests.Session] = None
) -> pd.DataFrame:
    """PNADC table 4092 (IBGE SIDRA): population 14+, labour force, employed, unemployed."""  # noqa: E501
    s = session or make_session()
    rows: List[dict] = []
    for p in periods:
        url = SIDRA.format(period=f"{p.year}{p.quarter:02d}")
        data = s.get(url, timeout=120).json()[1:]
        counts = {}
        for r in data:
            if r["D2N"].startswith("Pessoas de 14 anos") and r["MN"] == "Mil pessoas":
                counts[r["D4N"]] = float(r["V"].replace(",", ".")) * 1000
        rows += _rows(
            p,
            url,
            counts["Total"],
            counts["Força de trabalho"],
            counts["Força de trabalho - ocupada"],
            counts["Força de trabalho - desocupada"],
            SIDRA_TOL,
            "14+",
        )
    return pd.DataFrame(rows)


# ---------------------------------------------------------------- Colombia
DANE_ANNEX = "https://www.dane.gov.co/files/operaciones/GEIH/anex-GEIH-{mmm}{yyyy}.xlsx"
DANE_MONTHS = [
    "ene",
    "feb",
    "mar",
    "abr",
    "may",
    "jun",
    "jul",
    "ago",
    "sep",
    "oct",
    "nov",
    "dic",
]
DANE_QUARTERS = {"Ene - Mar": 1, "Abr - Jun": 2, "Jul - Sep": 3, "Oct - Dic": 4}
DANE_TOL = 0.05


def _dane_block(ws) -> Dict[str, list]:
    """First (national) block of a DANE annex sheet: year/period headers and label rows."""  # noqa: E501
    rows = list(ws.iter_rows(values_only=True))
    i = next(j for j, r in enumerate(rows) if r and r[0] == "Concepto")
    years, cur = [], None
    for y in rows[i][1:]:
        if y is not None:
            cur = str(y)[:4]
        years.append(cur)
    labels = {}
    for r in rows[i + 2 :]:
        if r and r[0] is not None and str(r[0]).startswith("Total "):
            break
        if r and r[0] is not None:
            labels[str(r[0]).strip().lower()] = r
    return {"years": years, "periods": list(rows[i + 1][1:]), "labels": labels}


def _dane_value(block: dict, prefix: str, col: int) -> float:
    key = next(k for k in block["labels"] if k.startswith(prefix))
    return float(block["labels"][key][col + 1])


def latest_dane_annex(session: requests.Session, newest: Period) -> Path:
    """Download the most recent 'anex-GEIH-<mmm><yyyy>.xlsx' that exists."""
    candidates = [Period(f"{newest.year + 1}M{m:02d}") for m in range(12, 0, -1)]
    candidates += [Period(f"{newest.year}M{m:02d}") for m in range(12, 0, -1)]
    for cand in candidates:
        url = DANE_ANNEX.format(mmm=DANE_MONTHS[cand.month - 1], yyyy=cand.year)
        if url_exists(url, session):
            res = download(
                url, RAW / "col" / "geih" / "docs" / Path(url).name, session=session
            )
            if res.status != "failed":
                return res.path
    raise FileNotFoundError("No DANE annex workbook found")


def annex_col(
    periods: Iterable[Period],
    annex_path: Optional[Path] = None,
    session: Optional[requests.Session] = None,
) -> pd.DataFrame:
    """DANE 'anexo GEIH' workbook, national calendar quarters (sheet 'Total nacional Trim')."""  # noqa: E501
    periods = list(periods)
    if annex_path is None:
        annex_path = latest_dane_annex(session or make_session(), max(periods))
    wb = openpyxl.load_workbook(annex_path, read_only=True, data_only=True)
    block = _dane_block(wb["Total nacional Trim"])
    src = f"DANE anexo GEIH {annex_path.name}, sheet 'Total nacional Trim'"
    wanted = {str(p): p for p in periods}
    rows: List[dict] = []
    for col, (year, label) in enumerate(zip(block["years"], block["periods"])):
        if year is None or label not in DANE_QUARTERS:
            continue
        key = f"{year}Q{DANE_QUARTERS[label]}"
        if key not in wanted:
            continue
        pet = _dane_value(block, "población en edad de trabajar", col) * 1000
        lf = _dane_value(block, "fuerza de trabajo", col) * 1000
        emp = _dane_value(block, "población ocupada", col) * 1000
        unemp = _dane_value(block, "población desocupada", col) * 1000
        rows += _rows(wanted[key], src, pet, lf, emp, unemp, DANE_TOL, "15+")
    return pd.DataFrame(rows)


# ---------------------------------------------------------------- Mexico
INEGI_BULLETINS = "https://www.inegi.org.mx/contenidos/saladeprensa/boletines/{path}"
# Quarterly ENOE press bulletins under the boletines folder (published two
# months after the quarter). Earlier bulletins use other folders and layouts;
# add paths here as they are located.
BULLETIN_PATHS = {
    "2023Q4": "2024/enoe/enoe2024_02.pdf",
    "2024Q1": "2024/enoe/enoe2024_05.pdf",
    "2024Q4": "2025/enoe/enoe2025_02.pdf",
    "2025Q1": "2025/enoe/enoe2025_05.pdf",
    "2025Q2": "2025/enoe/enoe2025_08.pdf",
    "2025Q3": "2025/enoe/enoe2025_11.pdf",
    "2025Q4": "2026/enoe/enoe2026_02.pdf",
    "2026Q1": "2026/enoe/enoe2026_05.pdf",
    "2026Q2": "2026/enoe/enoe2026_08.pdf",
}
INEGI_TOL = 0.06  # bulletin rates are rounded to one decimal


def parse_bulletin_text(text: str) -> Dict[str, float]:
    """Headline figures from an ENOE quarterly bulletin's text (population 15+)."""
    t = re.sub(r"\s+", " ", text)
    out: Dict[str, float] = {}
    m = re.search(r"tasa de desocupaci[oó]n[^.%]{0,60}?(\d{1,2}\.\d) ?%", t, re.I)
    if m:
        out["unemployment_rate"] = float(m.group(1))
    m = re.search(r"tasa de participaci[oó]n[^.%]{0,160}?(\d{2}\.\d) ?%", t, re.I)
    if m:
        out["participation_rate"] = float(m.group(1))
    m = re.search(
        r"(\d{2}\.\d) millones de personas estuvieron ocupadas"
        r"|poblaci[oó]n ocupada[^.]{0,40}?(\d{2}\.\d) millones",
        t,
        re.I,
    )
    if m:
        out["employed_millions"] = float(next(g for g in m.groups() if g))
    return out


def bulletins_mex(
    periods: Iterable[Period], session: Optional[requests.Session] = None
) -> pd.DataFrame:
    """INEGI quarterly ENOE bulletins: participation, unemployment, employed (15+)."""
    import pypdf

    s = session or make_session()
    rows: List[dict] = []
    for p in periods:
        path = BULLETIN_PATHS.get(str(p))
        if not path:
            continue
        url = INEGI_BULLETINS.format(path=path)
        dest = RAW / "mex" / "enoe" / "docs" / "bulletins" / Path(path).name
        res = download(url, dest, session=s)
        if res.status == "failed":
            continue
        reader = pypdf.PdfReader(res.path)
        text = " ".join((pg.extract_text() or "") for pg in reader.pages[:4])
        figs = parse_bulletin_text(text)
        base = {
            "period": str(p),
            "population": "15+",
            "source_url": url,
            "tolerance": INEGI_TOL,
        }
        for ind in ("participation_rate", "unemployment_rate"):
            if ind in figs:
                rows.append(dict(base, indicator=ind, value=figs[ind]))
        if "employed_millions" in figs:
            emp = figs["employed_millions"] * 1e6
            rows.append(dict(base, indicator="employed", value=emp, tolerance=0.05e6))
    return pd.DataFrame(rows)


FETCHERS = {"bra": sidra_bra, "col": annex_col, "mex": bulletins_mex}


# ---------------------------------------------------------------- Argentina
DATOS_AR = "https://apis.datos.gob.ar/series/api/series/?ids={ids}&start_date={start}&format=json"
ARG_SERIES = {
    "population": "49.2_TAEP_0_0_37",  # población total, 31 aglomerados, miles
    "labour_force": "49.2_TAEP_0_0_25",  # población económicamente activa, miles
    "employed": "49.2_TAEO_0_0_30",  # población ocupada, miles
}
ARG_TOL = 0.06  # INDEC publishes rates to one decimal; counts rounded to thousands


def series_arg(
    periods: Iterable[Period], session: Optional[requests.Session] = None
) -> pd.DataFrame:
    """INDEC EPH totals (31 urban agglomerations) from Argentina's series API.

    Rates use the whole population as the base, as INDEC does, so the rows
    carry ``population = "all"``.
    """
    periods = list(periods)
    s = session or make_session()
    ids = ",".join(ARG_SERIES.values())
    url = DATOS_AR.format(ids=ids, start=f"{min(periods).year}-01-01")
    data = s.get(url, timeout=120).json()
    order = [m["field"]["id"] for m in data["meta"][1:]]
    rows: List[dict] = []
    wanted = {str(p): p for p in periods}
    for rec in data["data"]:
        date = rec[0]
        year, month = int(date[:4]), int(date[5:7])
        key = f"{year}Q{(month - 1) // 3 + 1}"
        if key not in wanted or any(v is None for v in rec[1:]):
            continue
        vals = dict(zip(order, rec[1:]))
        total = vals[ARG_SERIES["population"]] * 1000
        lf = vals[ARG_SERIES["labour_force"]] * 1000
        emp = vals[ARG_SERIES["employed"]] * 1000
        rows += _rows(wanted[key], url, total, lf, emp, lf - emp, ARG_TOL, "all")
    return pd.DataFrame(rows)


# ---------------------------------------------------------------- Ecuador
INEC_TAB = (
    "https://www.ecuadorencifras.gob.ec/documentos/web-inec/EMPLEO/{y}/Trimestre_{r}/"
    "{y}_{r}_trimestre_Tabulados_Mercado_Laboral.xlsx"
)
INEC_TOL = 0.05
_ROMAN = {"I": 1, "II": 2, "III": 3, "IV": 4}


def _previous_quarter(p: Period) -> Period:
    return (
        Period(f"{p.year - 1}Q4")
        if p.quarter == 1
        else Period(f"{p.year}Q{p.quarter - 1}")
    )


def latest_inec_tabulados(session: requests.Session, newest: Period) -> Path:
    """Download the newest quarterly 'Tabulados Mercado Laboral' workbook."""
    cand = Period(f"{newest.year}Q{newest.quarter}")
    for _ in range(8):
        url = INEC_TAB.format(y=cand.year, r=cand.roman)
        if url_exists(url, session):
            res = download(
                url, RAW / "ecu" / "enemdu" / "docs" / Path(url).name, session=session
            )
            if res.status != "failed":
                return res.path
        cand = _previous_quarter(cand)
    raise FileNotFoundError("No INEC tabulados workbook found")


def tabulados_ecu(
    periods: Iterable[Period],
    tab_path: Optional[Path] = None,
    session: Optional[requests.Session] = None,
) -> pd.DataFrame:
    """INEC ENEMDU tabulados, sheet '1. Poblaciones', national column (15+)."""
    periods = list(periods)
    if tab_path is None:
        tab_path = latest_inec_tabulados(session or make_session(), max(periods))
    wb = openpyxl.load_workbook(tab_path, read_only=True, data_only=True)
    ws = wb["1. Poblaciones"]
    counts: Dict[str, Dict[str, float]] = {}
    for r in ws.iter_rows(values_only=True):
        if not r or r[0] is None or r[1] is None or not isinstance(r[2], (int, float)):
            continue
        label = re.sub(r"\s+", "", str(r[0]))  # ' IV - 2020' -> 'IV-2020'
        m = re.match(r"^(I|II|III|IV)-(\d{4})$", label)
        if not m:
            continue
        key = f"{m.group(2)}Q{_ROMAN[m.group(1)]}"
        counts.setdefault(key, {})[str(r[1]).strip()] = float(r[2])
    src = f"INEC tabulados {tab_path.name}, sheet '1. Poblaciones'"
    rows: List[dict] = []
    for p in periods:
        c = counts.get(str(p))
        if not c:
            continue
        pet = next(
            v for k, v in c.items() if k.startswith("Población en Edad de Trabajar")
        )
        pea = next(
            v for k, v in c.items() if k.startswith("Población Económicamente Activa")
        )
        rows += _rows(p, src, pet, pea, c["Empleo"], c["Desempleo"], INEC_TOL, "15+")
    return pd.DataFrame(rows)


FETCHERS.update({"arg": series_arg, "ecu": tabulados_ecu})


# ---------------------------------------------------------------- Peru
# INEI "Comportamiento de los indicadores del mercado laboral a nivel nacional
# y 27 ciudades": quarterly PDF reports published through gob.pe. Each report
# tabulates the quarter and the same quarter a year earlier, so the 2023
# reports also cover 2022. The gob.pe id of each report is recorded by hand.
GOBPE_PAGES = {
    "2023Q1": 4519379, "2023Q2": 4640942, "2023Q3": 4848417, "2023Q4": 5196099,
    "2024Q1": 5575916, "2024Q2": 5874417, "2024Q3": 6188990, "2024Q4": 6474256,
    "2025Q1": 6769752, "2025Q2": 7043994, "2025Q3": 7403694, "2025Q4": 7739601,
    "2026Q1": 8134830, "2026Q2": 8484921,
}  # fmt: skip
GOBPE_PAGE = "https://www.gob.pe/institucion/inei/informes-publicaciones/{id}"
GOBPE_HEADERS = {  # gob.pe answers 418 to non-browser clients
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/126.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/pdf;q=0.9,*/*;q=0.8",
    "Accept-Language": "es-PE,es;q=0.9",
}
INEI_TOL = 0.05  # counts are published to the nearest hundred persons
_QUARTER_WORD = {
    "Primer": 1, "Segundo": 2, "Tercer": 3, "Cuarto": 4,
    "I": 1, "II": 2, "III": 3, "IV": 4,
}  # fmt: skip
_HEADER = (
    r"(Primer|Segundo|Tercer|Cuarto|I|II|III|IV)\s+[Tt]rimestre:?"
    r"\s*(\d{4})\s+y\s+(\d{4})"
)
_NUM = r"-?\s?\d{1,3}(?:\s\d{3})*,\d"
_PDF_LINK = r'https://cdn\.www\.gob\.pe/uploads/document/file/\d+/[^"\']+\.pdf'


def _inei_num(s: str) -> float:
    return float(s.replace(" ", "").replace(",", "."))


def _values(nums: List[str]) -> Optional[tuple]:
    if len(nums) >= 6:  # layout with a percentage column next to each year
        return _inei_num(nums[0]), _inei_num(nums[2])
    if len(nums) >= 2:
        return _inei_num(nums[0]), _inei_num(nums[1])
    return None


def _nearest(pattern: str, text: str, start: int, end: int, span: int = 1500):
    """Match of ``pattern`` nearest to the header at ``text[start:end]``.

    Reports up to 2025 print the rows after the table label and header;
    the 2025Q4 and 2026 layouts print them before, so the nearest match on
    either side is taken (the annual tables of the same page sit further away).
    """
    best = None
    lo, hi = max(0, start - span), min(len(text), end + span)
    for m in re.compile(pattern).finditer(text, lo, hi):
        dist = min(abs(m.start() - end), abs(m.end() - start))
        if best is None or dist < best[0]:
            best = (dist, m)
    return best[1] if best else None


_ROW = r"((?:\s+" + _NUM + r")+)"  # run of numbers after a row label
_PET_LABEL = "Población en Edad de Trabajar"
_PEA_LABEL = "Población Económicamente Activa"
_NATIONAL_BLOCK = r"Total\s*\n\s*" + _PET_LABEL + _ROW + r"\s*\n\s*" + _PEA_LABEL + _ROW


def _working_age_rows(text: str, start: int, end: int) -> Dict[str, tuple]:
    """National PET and PEA of table 1.1: the block headed ``Total`` where the
    layout has one (urban and rural blocks follow), else the rows nearest the
    header (2023 layout, no ``Total`` line)."""
    m = _nearest(_NATIONAL_BLOCK, text, start, end)
    if m:
        vals = {"pet": _values(re.findall(_NUM, m.group(1)))}
        vals["pea"] = _values(re.findall(_NUM, m.group(2)))
        return {k: v for k, v in vals.items() if v}
    out = {}
    for label, key in ((_PET_LABEL, "pet"), (_PEA_LABEL, "pea")):
        m = _nearest(re.escape(label) + _ROW, text, start, end)
        if m and _values(re.findall(_NUM, m.group(1))):
            out[key] = _values(re.findall(_NUM, m.group(1)))
    return out


def gobpe_pdf_url(page_id: int, session: requests.Session) -> str:
    """Resolve the CDN link of the PDF attached to a gob.pe publication page."""
    r = session.get(GOBPE_PAGE.format(id=page_id), headers=GOBPE_HEADERS, timeout=120)
    r.raise_for_status()
    m = re.search(
        r'https://cdn\.www\.gob\.pe/uploads/document/file/\d+/[^"\']+\.pdf', r.text
    )
    if not m:
        raise ValueError(f"No PDF link on gob.pe page {page_id}")
    return m.group(0)


def parse_inei_report(text: str) -> Dict[str, Dict[str, float]]:
    """Quarterly totals (thousands) from the report text: PET, PEA, employed.

    Tables carry a header like ``Primer trimestre: 2024 y 2025`` (Roman
    numerals from 2026) and two columns of absolute values, in some years
    each followed by a percentage column.
    """
    out: Dict[str, Dict[str, float]] = {}
    for m in re.finditer(_HEADER, text):
        q = _QUARTER_WORD[m.group(1)]
        keys = (f"{m.group(2)}Q{q}", f"{m.group(3)}Q{q}")
        context = text[max(0, m.start() - 300) : m.end() + 300].upper()
        found: Dict[str, tuple] = {}
        if "EDAD DE TRABAJAR" in context and "CONDICIÓN DE ACTIVIDAD" in context:
            found = _working_age_rows(text, m.start(), m.end())
        elif re.search(r"POBLACIÓN OCUPADA,? SEGÚN .REA DE RESIDENCIA", context):
            mm = _nearest(r"Total" + _ROW, text, m.start(), m.end())
            if mm and _values(re.findall(_NUM, mm.group(1))):
                found = {"emp": _values(re.findall(_NUM, mm.group(1)))}
        for key, vals in found.items():
            for k, v in zip(keys, vals):
                out.setdefault(k, {}).setdefault(key, v)
    return {k: v for k, v in out.items() if {"pet", "pea", "emp"} <= set(v)}


def informes_per(
    periods: Iterable[Period], session: Optional[requests.Session] = None
) -> pd.DataFrame:
    """INEI EPEN national totals (persons 14+) parsed from the quarterly reports."""
    from pypdf import PdfReader

    periods = list(periods)
    s = session or make_session()
    wanted = {str(p): p for p in periods}
    docs = RAW / "per" / "epen" / "docs"
    found: Dict[str, dict] = {}
    for key, page_id in GOBPE_PAGES.items():
        # a report covers its own quarter and the same quarter one year earlier
        prev = f"{int(key[:4]) - 1}{key[4:]}"
        if not ({key, prev} & set(wanted)):
            continue
        try:
            url = gobpe_pdf_url(page_id, s)
        except (requests.RequestException, ValueError) as exc:
            print(f"{key}: {exc}")
            continue
        res = download(url, docs / f"informe_{key}.pdf", session=s)
        if res.status == "failed":
            print(f"{key}: {res.error}")
            continue
        text = "\n".join(
            (pg.extract_text() or "") for pg in PdfReader(res.path).pages[:40]
        )
        for k, vals in parse_inei_report(text).items():
            found.setdefault(k, dict(vals, url=url))
    rows: List[dict] = []
    for key, p in wanted.items():
        v = found.get(key)
        if not v:
            continue
        pet, pea, emp = (1000 * v[x] for x in ("pet", "pea", "emp"))
        rows += _rows(p, v["url"], pet, pea, emp, pea - emp, INEI_TOL, "14+")
    return pd.DataFrame(rows)


FETCHERS.update({"per": informes_per})


# ---------------------------------------------------------------- South Africa
STATSSA_TRENDS = (
    "https://www.statssa.gov.za/publications/P0211/QLFS%20Trends%202008-{y}Q{q}.xlsx"
)
STATSSA_TOL = 0.05  # counts published in thousands
_QLFS_MONTHS = {"Jan-Mar": 1, "Apr-Jun": 2, "Jul-Sep": 3, "Oct-Dec": 4}


def latest_statssa_trends(session: requests.Session, newest: Period) -> Path:
    """Download the newest 'QLFS Trends 2008-<quarter>' workbook (Stats SA P0211)."""
    cand = Period(str(newest))
    for _ in range(8):
        url = STATSSA_TRENDS.format(y=cand.year, q=cand.quarter)
        if url_exists(url, session):
            res = download(
                url,
                RAW / "zaf" / "qlfs" / "docs" / Path(url).name.replace("%20", "_"),
                session=session,
            )
            if res.status != "failed":
                return res.path
        cand = _previous_quarter(cand)
    raise FileNotFoundError("No Stats SA QLFS Trends workbook found")


def trends_zaf(
    periods: Iterable[Period],
    tab_path: Optional[Path] = None,
    session: Optional[requests.Session] = None,
) -> pd.DataFrame:
    """Stats SA QLFS Trends, 'Table 2' (both sexes, persons 15-64, thousands)."""
    periods = list(periods)
    if tab_path is None:
        tab_path = latest_statssa_trends(session or make_session(), max(periods))
    wb = openpyxl.load_workbook(tab_path, read_only=True, data_only=True)
    ws = wb["Table 2"]
    rows = list(ws.iter_rows(values_only=True))

    def _is_quarter(v) -> bool:
        return isinstance(v, str) and v[:7] in _QLFS_MONTHS and v[8:].strip().isdigit()

    header = next(r for r in rows if _is_quarter(r[1]))
    cols = {}
    for i, h in enumerate(header):
        if _is_quarter(h):
            cols[f"{h[8:].strip()}Q{_QLFS_MONTHS[h[:7]]}"] = i
    wanted = {
        "Population 15-64": "pet",
        "Labour Force": "lf",
        "Employed": "emp",
        "Unemployed": "unemp",
    }
    values: Dict[str, Dict[str, float]] = {}
    seen = set()
    for r in rows:
        label = str(r[0]).strip() if r[0] else ""
        key = next((v for k, v in wanted.items() if label.startswith(k)), None)
        if key is None or key in seen:
            continue
        seen.add(key)  # first block is 'Both sexes'
        for period, i in cols.items():
            try:
                values.setdefault(period, {})[key] = 1000 * float(r[i])
            except (TypeError, ValueError):
                pass
    src = f"Stats SA {tab_path.name}, Table 2"
    out: List[dict] = []
    for p in periods:
        v = values.get(str(p))
        if v and {"pet", "lf", "emp", "unemp"} <= set(v):
            out += _rows(
                p, src, v["pet"], v["lf"], v["emp"], v["unemp"], STATSSA_TOL, "15-64"
            )
    return pd.DataFrame(out)


FETCHERS.update({"zaf": trends_zaf})
